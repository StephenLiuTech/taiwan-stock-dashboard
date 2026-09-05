"""Schema-v14 historical Stock Net Equity import regression tests."""

import sqlite3
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path

import pytest

from domain import (
    Currency,
    HistoricalStockNetEquity,
    Liability,
    LiabilityPrincipalEvent,
    LiabilityPrincipalEventType,
    LiabilityType,
    StockNetEquityQuality,
)
from pams.application import (
    ExcelV15StockNetEquitySource,
    ImportStockNetEquityHistoryUseCase,
    StockNetEquityImportError,
)
from pams.cli import build_parser
from repositories.market_data_uow import SQLiteStockNetEquityHistoryUnitOfWork
from repositories.sqlite import (
    SQLiteLiabilityPrincipalEventRepository,
    SQLiteLiabilityRepository,
    SQLiteSnapshotRepository,
    SQLiteStockNetEquityHistoryRepository,
)


class FixedSource:
    def __init__(self, rows: tuple[HistoricalStockNetEquity, ...]) -> None:
        self.rows = rows

    def read(self, path: Path) -> tuple[HistoricalStockNetEquity, ...]:
        del path
        return self.rows


def history_row(day: date = date(2023, 8, 7)) -> HistoricalStockNetEquity:
    return HistoricalStockNetEquity(
        snapshot_date=day,
        total_market_value=Decimal("81751.13"),
        pledge_debt=Decimal("0"),
        margin_debt=Decimal("0"),
        total_liabilities=Decimal("0"),
        stock_net_equity=Decimal("81751.13"),
        quality_status=StockNetEquityQuality.VERIFIED,
        source="fixture",
        source_reference="fixture:2",
        imported_at=datetime(2026, 9, 4, tzinfo=UTC),
    )


def use_case(
    connection: sqlite3.Connection,
    rows: tuple[HistoricalStockNetEquity, ...] | None = None,
    *,
    liabilities: list[Liability] | None = None,
    events: list[LiabilityPrincipalEvent] | None = None,
) -> tuple[
    ImportStockNetEquityHistoryUseCase,
    SQLiteStockNetEquityHistoryRepository,
]:
    if rows is None:
        rows = (history_row(),)
    repository = SQLiteStockNetEquityHistoryRepository(connection)
    return (
        ImportStockNetEquityHistoryUseCase(
            FixedSource(rows),
            repository,
            SQLiteSnapshotRepository(connection),
            (
                liabilities
                if liabilities is not None
                else SQLiteLiabilityRepository(connection).list_all()
            ),
            (
                events
                if events is not None
                else SQLiteLiabilityPrincipalEventRepository(connection).list_all()
            ),
            SQLiteStockNetEquityHistoryUnitOfWork(connection),
            "test",
        ),
        repository,
    )


def liability_fixture() -> list[Liability]:
    return [
        Liability(
            id="pledge",
            liability_type=LiabilityType.STOCK_PLEDGE,
            principal=Decimal("998000"),
            currency=Currency.TWD,
        ),
        Liability(
            id="margin",
            liability_type=LiabilityType.MARGIN_FINANCING,
            principal=Decimal("543000"),
            currency=Currency.TWD,
        ),
    ]


def principal_event(
    event_id: str,
    liability_id: str,
    day: date,
    delta: str,
    resulting: str,
) -> LiabilityPrincipalEvent:
    return LiabilityPrincipalEvent(
        id=event_id,
        liability_id=liability_id,
        effective_date=day,
        sequence=10,
        event_type=LiabilityPrincipalEventType.INCREASE,
        principal_delta=Decimal(delta),
        resulting_principal=Decimal(resulting),
        source="fixture",
    )


def modeled_row(day: date, market: str, net_equity: str) -> HistoricalStockNetEquity:
    return HistoricalStockNetEquity(
        snapshot_date=day,
        total_market_value=Decimal(market),
        pledge_debt=Decimal("481855.2327221438"),
        margin_debt=Decimal("0"),
        total_liabilities=Decimal("481855.2327221438"),
        stock_net_equity=Decimal(net_equity),
        quality_status=StockNetEquityQuality.ESTIMATED_LIABILITY,
        source="股票歷史資產.xlsx",
        source_reference="workbook estimate",
    )


def test_default_cli_mode_is_dry_run_and_apply_is_explicit() -> None:
    parser = build_parser()
    default = parser.parse_args(
        ["stock-net-equity", "import", "--source", "history.xlsx"]
    )
    applied = parser.parse_args(
        ["stock-net-equity", "import", "--source", "history.xlsx", "--apply"]
    )

    assert default.apply is False
    assert applied.apply is True


def test_dry_run_does_not_write(connection: sqlite3.Connection) -> None:
    case, repository = use_case(connection)

    result = case.execute(Path("ignored.xlsx"), apply=False)

    assert result.missing == 1
    assert result.inserted == 0
    assert repository.list_between_dates(date.min, date.max) == []


def test_apply_is_insert_only_and_idempotent(connection: sqlite3.Connection) -> None:
    case, repository = use_case(connection)

    first = case.execute(Path("ignored.xlsx"), apply=True)
    second = case.execute(Path("ignored.xlsx"), apply=True)

    assert first.inserted == 1
    assert second.inserted == 0
    assert repository.list_between_dates(date.min, date.max) == [history_row()]


def test_import_does_not_create_absent_weekend_rows(
    connection: sqlite3.Connection,
) -> None:
    friday = history_row(date(2026, 8, 7))
    monday = history_row(date(2026, 8, 10))
    case, repository = use_case(connection, (friday, monday))

    result = case.execute(Path("ignored.xlsx"), apply=True)

    assert result.inserted == 2
    assert [
        item.snapshot_date for item in repository.list_between_dates(date.min, date.max)
    ] == [
        date(2026, 8, 7),
        date(2026, 8, 10),
    ]


def test_unknown_quality_allows_auditable_gap_without_zero(
    connection: sqlite3.Connection,
) -> None:
    unknown = HistoricalStockNetEquity(
        snapshot_date=date(2026, 8, 10),
        quality_status=StockNetEquityQuality.UNKNOWN,
        source="production reconciliation",
        source_reference="liability mismatch",
    )
    case, repository = use_case(connection, (unknown,))

    case.execute(Path("ignored.xlsx"), apply=True)

    persisted = repository.list_between_dates(date.min, date.max)[0]
    assert persisted.stock_net_equity is None
    assert persisted.quality_status is StockNetEquityQuality.UNKNOWN


def test_excel_v15_source_maps_controlled_quality_and_decimal_values(
    tmp_path: Path,
) -> None:
    import openpyxl

    path = tmp_path / "v15.xlsx"
    workbook = openpyxl.Workbook()
    assets = workbook.active
    assets.title = "資產趨勢"
    assets.append(["日期", "B", "C", "D", "E", "F", "淨值", "總市值"])
    assets.append([date(2023, 8, 7), 1, 1, 0, 0, 0, 81751.13, 81751.13])
    liabilities = workbook.create_sheet("負債資料")
    liabilities.append(["日期", "質押", "融資", "總負債", "品質"])
    liabilities.append([date(2023, 8, 7), 0, 0, 0, "無負債"])
    prices = workbook.create_sheet("每日價格")
    prices.append(["日期", "2330"])
    prices.append([date(2023, 8, 7), 558])
    prices.append([date(2023, 8, 8), 558])
    workbook.save(path)

    rows = ExcelV15StockNetEquitySource().read(path)

    assert len(rows) == 1079
    assert rows[0].snapshot_date == date(2023, 8, 7)
    assert rows[0].stock_net_equity == Decimal("81751.13")
    assert rows[0].quality_status is StockNetEquityQuality.VERIFIED
    assert rows[1].quality_status is StockNetEquityQuality.UNKNOWN


def test_excel_v15_marks_carried_forward_prices(tmp_path: Path) -> None:
    import openpyxl

    path = tmp_path / "v15.xlsx"
    workbook = openpyxl.Workbook()
    assets = workbook.active
    assets.title = "資產趨勢"
    assets.append(["日期", "B", "C", "D", "E", "F", "淨值", "總市值"])
    liabilities = workbook.create_sheet("負債資料")
    liabilities.append(["日期", "質押", "融資", "總負債", "品質"])
    prices = workbook.create_sheet("每日價格")
    prices.append(["日期", "2330"])
    for day in (date(2023, 8, 7), date(2023, 8, 8)):
        assets.append([day, 1, 1, 0, 0, 0, 100, 100])
        liabilities.append([day, 0, 0, 0, "無負債"])
        prices.append([day, 558])
    workbook.save(path)

    rows = ExcelV15StockNetEquitySource().read(path)

    assert "price=carried-forward" not in rows[0].source_reference
    assert "price=carried-forward" in rows[1].source_reference


def test_importer_rebuilds_modeled_liabilities_from_as_of_event_ledger(
    connection: sqlite3.Connection,
) -> None:
    rows = (
        modeled_row(date(2026, 1, 6), "1488340", "1006484.7672778561"),
        modeled_row(date(2026, 5, 15), "2367950", "1886094.7672778562"),
        modeled_row(date(2026, 6, 2), "3924775", "3442919.7672778562"),
        modeled_row(date(2026, 7, 20), "3826460", "3344604.7672778562"),
    )
    events = [
        principal_event("p-1", "pledge", date(2026, 1, 6), "483000", "483000"),
        principal_event("p-2", "pledge", date(2026, 5, 15), "455000", "938000"),
        principal_event("p-3", "pledge", date(2026, 5, 25), "60000", "998000"),
        principal_event("m-1", "margin", date(2026, 6, 2), "50000", "50000"),
        principal_event("m-2", "margin", date(2026, 7, 14), "493000", "543000"),
        # A future omission/conflict cannot invalidate an earlier as-of prefix.
        principal_event("m-future", "margin", date(2026, 8, 31), "57720", "1"),
    ]
    case, _ = use_case(connection, rows, liabilities=liability_fixture(), events=events)

    result = case.execute(Path("股票歷史資產.xlsx"), apply=False)
    by_date = {row.snapshot_date: row for row in result.rows}

    assert by_date[date(2026, 1, 6)].stock_net_equity == Decimal("1005340")
    assert by_date[date(2026, 5, 15)].stock_net_equity == Decimal("1429950")
    assert by_date[date(2026, 6, 2)].stock_net_equity == Decimal("2876775")
    assert by_date[date(2026, 7, 20)].stock_net_equity == Decimal("2285460")
    assert all(
        row.quality_status is StockNetEquityQuality.VERIFIED for row in result.rows
    )
    assert all(
        "market_value=股票歷史資產.xlsx:資產趨勢!H" in row.source_reference
        and "liabilities=production_event_ledger_as_of_date" in row.source_reference
        for row in result.rows
    )


def test_importer_stops_when_event_history_has_no_zero_based_opening(
    connection: sqlite3.Connection,
) -> None:
    rows = (modeled_row(date(2026, 1, 6), "1488340", "1006484.7672778561"),)
    events = [
        principal_event("p-1", "pledge", date(2026, 1, 6), "83000", "483000"),
        principal_event("m-1", "margin", date(2026, 6, 2), "50000", "50000"),
    ]
    case, _ = use_case(connection, rows, liabilities=liability_fixture(), events=events)

    with pytest.raises(StockNetEquityImportError, match="zero-based opening"):
        case.execute(Path("股票歷史資產.xlsx"), apply=False)


def test_importer_rejects_existing_rows_with_different_rebuilt_values(
    connection: sqlite3.Connection,
) -> None:
    existing = modeled_row(date(2026, 1, 6), "1488340", "1006484.7672778561")
    SQLiteStockNetEquityHistoryRepository(connection).insert_many_if_absent([existing])
    events = [
        principal_event("p-1", "pledge", date(2026, 1, 6), "483000", "483000"),
        principal_event("m-1", "margin", date(2026, 6, 2), "50000", "50000"),
    ]
    case, _ = use_case(
        connection,
        (existing,),
        liabilities=liability_fixture(),
        events=events,
    )

    with pytest.raises(StockNetEquityImportError, match="conflicts"):
        case.execute(Path("股票歷史資產.xlsx"), apply=True)
